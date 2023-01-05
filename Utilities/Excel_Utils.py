import openpyxl
import pandas as pd

#df = pd.DataFrame()
#df.read_csv(".\\Test_Data\\traveller_information.csv")


#print(df)
class get_traveller_info:

    def read_csv(self):
        try:
            #df = pd.read_csv('Nifty_50.csv')
            #df = pd.DataFrame()
            #df = df.reindex(columns=['Symbol','LastRate','High','Low','PClose','ChgPcnt','Chg','TotalQty','Time'])
            df = pd.read_csv('Test_Data/traveller_information.csv')
            print(df)
            x = df.shape
            counter = len(df)
            print(counter)
            print(x)
            rowLabels = df.index.tolist()
            col_names=df.columns.values
            print(col_names)
        except Exception as ex:
            print(ex)


get_traveller_info().read_csv()
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=row_num, column=column_num).value


def writeData(file, sheetName, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)
